import pandas as pd
import os
from openpyxl.utils import get_column_letter

def generate_report(df_invoice: pd.DataFrame,
                    df_verification: pd.DataFrame,
                    output_path: str) -> None:
    """
    Creates a multi-sheet Excel report with:
    - Extracted line items (with FX conversion)
    - Verification results
    - Summary sheet with calculated KPIs
    """

    # Ensure output folder exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # 1️⃣ Write the invoice extraction sheet
        df_invoice.to_excel(writer, sheet_name="Extracted Line Items", index=False)

        # Auto-size columns
        sheet = writer.sheets["Extracted Line Items"]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(60, length + 2)

        # 2️⃣ Write the verification sheet
        df_verification.to_excel(writer, sheet_name="Verification Results", index=False)

        sheet = writer.sheets["Verification Results"]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(60, length + 2)

        # 3️⃣ Build a summary sheet
        summary_data = {
            "total_items_in_invoice": len(df_invoice),
            "items_with_supporting_docs": int(df_verification["supporting_attached"].sum()),
            "items_with_matching_values": int(df_verification["non_vat_match"].sum()),
            "items_missing_supporting_docs": int(len(df_verification) - df_verification["supporting_attached"].sum()),
            "items_not_matching_value": int(len(df_verification) - df_verification["non_vat_match"].sum()),
            "total_converted_non_vat_value": float(df_invoice["converted_non_vat_value"].fillna(0).sum()),
        }

        df_summary = pd.DataFrame([summary_data])
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        sheet = writer.sheets["Summary"]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(60, length + 2)

    print(f"[INFO] Excel validation report saved to: {output_path}")
